"""Rate-table operations (03 §3.3, slices W10-2 and W10-3B): seeding, validation,
cell diffs, the four bulk operations of FR-RATE-18, and the CSV/XLSX import/export
of FR-RATE-20.

The pricing actions behind FR-RATE-16 (seed a table from an approved model),
FR-RATE-19 (validation on save, checked before the version persists) and FR-RATE-17
(cell-level diffs with exposure weighting). Pure operations: cells are rows of
decimal strings (R2), exposure weights are supplied by the caller at diff-fetch time
(DP1), and no persistence happens here. Each bulk operation returns a new immutable
version carrying its `BulkOperation` record (04 §4.4), and imports preview the
would-be version's diff against the addressed version (03 §5.2) before anything
persists.
"""

from __future__ import annotations

import csv
import hashlib
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from itertools import product
from typing import Literal
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from model_schema.modelling import GlmFitResult, Model, RelativityLevel
from model_schema.rating import (
    BulkOperation,
    BulkOperationResult,
    FloorAndCapOperation,
    FloorAndCapParameters,
    ImportPreview,
    ImportVerdict,
    KeyFilter,
    RateTable,
    RateTableDiff,
    RateTableKey,
    RateTableKeyType,
    RateTableStorageMode,
    RateTableValue,
    RateTableValueType,
    RateTableVersion,
    RebaseToLevelOperation,
    RebaseToLevelParameters,
    SeededFrom,
    UpliftByFilterOperation,
    UpliftByFilterParameters,
    UpliftTableOperation,
    UpliftTableParameters,
)
from model_schema.refs import ArtifactRef
from pricing_core.rating.compile import ValidationIssue

_APPROVED_OR_BETTER = frozenset({"approved", "live", "retired"})

#: The plan's named validation failure codes (plan §Slice W10-2 T3). Internal codes:
#: the API maps them onto 03 §5.1's module error codes at the boundary.
INCOMPLETE_KEY_DOMAIN = "INCOMPLETE_KEY_DOMAIN"
NULL_VALUE = "NULL_VALUE"
OUT_OF_BOUNDS = "OUT_OF_BOUNDS"
DUPLICATE_KEY = "DUPLICATE_KEY"

#: W10-3B's named failure codes (03 §5.2, FR-RATE-18/20/62). Internal codes: the API
#: maps them onto 03 §5.1's module error codes at the boundary.
PARQUET_CELLS_UNAVAILABLE = "PARQUET_CELLS_UNAVAILABLE"
FILTER_UNKNOWN_KEY = "FILTER_UNKNOWN_KEY"
FLOOR_ABOVE_CAP = "FLOOR_ABOVE_CAP"
REBASE_NO_MATCH = "REBASE_NO_MATCH"
REBASE_AMBIGUOUS = "REBASE_AMBIGUOUS"
REBASE_ZERO_REFERENCE = "REBASE_ZERO_REFERENCE"
IMPORT_KEY_MISMATCH = "IMPORT_KEY_MISMATCH"
IMPORT_TYPE_MISMATCH = "IMPORT_TYPE_MISMATCH"
IMPORT_PARSE_ERROR = "IMPORT_PARSE_ERROR"

#: One cell row: the declared key columns plus the value column, all decimal strings.
CellRow = dict[str, str]
Cells = Sequence[CellRow]

#: The key of a row: one value per declared key column, in key-declaration order.
KeyTuple = tuple[str, ...]

#: Per-cell exposure weights, keyed by row key (DP1: supplied at diff-fetch time).
Weights = Mapping[KeyTuple, Decimal | int | float]


@dataclass(frozen=True)
class SeedResult:
    """What seeding produces: the table definition, the cells, and the seed origin."""

    table: RateTable
    cells: tuple[CellRow, ...]
    seeded_from: SeededFrom


def check_model_approved(model: Model) -> None:
    """FR-OVR-14: only an approved (or better) model may seed a rate table."""
    if model.status not in _APPROVED_OR_BETTER:
        raise ValueError(
            f"PIN_NOT_APPROVED: model {model.model_family_slug}@{model.version} "
            f"is {model.status}, not approved"
        )


def _glm_relativities(
    model: Model,
) -> dict[str, tuple[RelativityLevel, ...]]:
    """The GLM relativity table, or a named refusal for any other fit kind.

    Only `GlmFitResult` carries per-level relativities (ADR-0003): a GBM or EBM fit has
    nothing to seed a rate table from. Raised as a plain message rather than a coded
    one so the API's boundary maps it to `VALIDATION_FAILED` — the code shape is
    reserved for codes the owning spec enumerates.
    """
    fit = model.fit_result
    if not isinstance(fit, GlmFitResult):
        kind = fit.model_type if fit is not None else "absent"
        raise ValueError(
            f"model {model.model_family_slug}@{model.version} carries a {kind} fit "
            "result; rate-table seeding reads GLM relativities"
        )
    return fit.relativities


def extract_relativity_table(
    model: Model, *, value_name: str = "relativity"
) -> list[CellRow]:
    """The model's factor relativities as cell rows (FR-RATE-16).

    Levels without a relativity (non-log links) are skipped: they have nothing to
    seed, and seeding 1.0 there would fabricate a technical rate. Values are
    rendered as decimal strings, never JSON floats (R2).
    """
    rows: list[CellRow] = []
    for factor, levels in _glm_relativities(model).items():
        for level in levels:
            if level.relativity is None:
                continue
            rows.append({factor: level.level, value_name: str(Decimal(str(level.relativity)))})
    return rows


def _key_domains_of(model: Model) -> dict[str, frozenset[str]]:
    """One domain per factor: the levels that carry a seeded relativity."""
    return {
        factor: frozenset(level.level for level in levels if level.relativity is not None)
        for factor, levels in _glm_relativities(model).items()
    }


def seed_from_model(
    model: Model,
    *,
    table_slug: str,
    change_note: str,
    seeded_at: datetime,
    rateable: bool = True,
    value_name: str = "relativity",
) -> SeedResult:
    """Seed a rate table from an approved model (FR-RATE-16, plan §Slice W10-2 T1).

    Builds the table definition from the model's factor relativities, validates the
    extracted cells (T3's check runs before the version persists), and pins the seed
    origin so "how far from the technical rate?" is answerable.
    """
    check_model_approved(model)
    cells = tuple(extract_relativity_table(model, value_name=value_name))
    if not cells:
        raise ValueError(
            f"NO_RELATIVITIES: model {model.model_family_slug}@{model.version} "
            "carries no relativities"
        )

    keys = [
        RateTableKey(name=factor, type=RateTableKeyType.STRING, banding_ref=None)
        for factor in _glm_relativities(model)
    ]
    value = RateTableValue(
        name=value_name,
        type=RateTableValueType.RELATIVITY,
        unit="factor",
        min=None,
        max=None,
    )
    issues = validate_rate_table(cells, keys, value, key_domains=_key_domains_of(model))
    if issues:
        raise ValueError(f"{issues[0].code}: {issues[0].message}")

    table = RateTable(
        slug=table_slug,
        version=1,
        rateable=rateable,
        storage=RateTableStorageMode.ROWS,
        keys=keys,
        value=value,
        default_row=None,
    )
    seeded_from = SeededFrom(
        model_ref=ArtifactRef(
            type="model", slug=model.model_family_slug, version=model.version
        ),
        seeded_at=seeded_at,
    )
    return SeedResult(table=table, cells=cells, seeded_from=seeded_from)


def _value_issue(
    row: CellRow, value: RateTableValue, key_names: Sequence[str]
) -> ValidationIssue | None:
    """The value half of FR-RATE-19 for one row: non-null, decimal, in bounds.

    Shared by save-time validation and by the bulk operations' result check (03 §5.2:
    each operation validates the result before persisting).
    """
    key_values = tuple(row[key] for key in key_names)
    raw = row.get(value.name)
    if raw is None or raw == "":
        return ValidationIssue(
            code=NULL_VALUE,
            message=f"no value for key {key_values}",
            field=value.name,
        )
    try:
        decimal_value = Decimal(raw)
    except InvalidOperation:
        return ValidationIssue(
            code=OUT_OF_BOUNDS,
            message=f"value {raw!r} for key {key_values} is not a decimal number",
            field=value.name,
        )
    if value.min is not None and decimal_value < value.min:
        return ValidationIssue(
            code=OUT_OF_BOUNDS,
            message=(
                f"value {raw} for key {key_values} is below the declared "
                f"minimum {value.min}"
            ),
            field=value.name,
        )
    if value.max is not None and decimal_value > value.max:
        return ValidationIssue(
            code=OUT_OF_BOUNDS,
            message=(
                f"value {raw} for key {key_values} is above the declared "
                f"maximum {value.max}"
            ),
            field=value.name,
        )
    return None


def validate_rate_table(
    cells: Cells,
    keys: Sequence[RateTableKey],
    value: RateTableValue,
    *,
    key_domains: Mapping[str, frozenset[str]],
    default_row: CellRow | None = None,
) -> list[ValidationIssue]:
    """FR-RATE-19: named validation, checked before the version persists.

    Coverage is the Cartesian product of the declared key domains unless an explicit
    `default_row` waives it. Every row must have a non-null, in-bounds decimal value,
    and no duplicate keys. Failures are named per plan T3: INCOMPLETE_KEY_DOMAIN,
    NULL_VALUE, OUT_OF_BOUNDS, DUPLICATE_KEY.
    """
    issues: list[ValidationIssue] = []
    key_names = [key.name for key in keys]

    seen: set[KeyTuple] = set()
    for row in cells:
        key_values = tuple(row[key] for key in key_names)
        if key_values in seen:
            issues.append(
                ValidationIssue(
                    code=DUPLICATE_KEY,
                    message=f"key {key_values} appears more than once",
                    field=".".join(key_names),
                )
            )
            continue
        seen.add(key_values)

        issue = _value_issue(row, value, key_names)
        if issue is not None:
            issues.append(issue)

    if default_row is None:
        missing = [
            combination
            for combination in product(*(key_domains[key] for key in key_names))
            if combination not in seen
        ]
        for combination in missing:
            issues.append(
                ValidationIssue(
                    code=INCOMPLETE_KEY_DOMAIN,
                    message=f"missing value for key combination {combination}",
                    field=".".join(key_names),
                )
            )
    return issues


def _index_rows(
    cells: Cells, key_names: Sequence[str], value_name: str
) -> dict[KeyTuple, Decimal]:
    """Map each row's key to its value. Callers hand validated cells (unique keys)."""
    return {
        tuple(row[key] for key in key_names): Decimal(row[value_name]) for row in cells
    }


def _compute_diff(
    baseline: Cells,
    current: Cells,
    keys: Sequence[RateTableKey],
    value: RateTableValue,
    weights: Weights | None,
) -> RateTableDiff:
    """The shared core of diff_vs_previous and diff_vs_seed (FR-RATE-17).

    A cell counts as changed when its value differs from the baseline; an added or
    removed key counts. Percentage statistics cover cells comparable in both
    directions with a non-zero baseline, so a cell born from or into zero never
    fabricates an infinite change. The exposure-weighted mean covers the comparable
    cells that carry a weight (DP1: weights at fetch time).
    """
    key_names = [key.name for key in keys]
    before = _index_rows(baseline, key_names, value.name)
    after = _index_rows(current, key_names, value.name)

    changed = sorted(
        key for key in before.keys() | after.keys() if before.get(key) != after.get(key)
    )
    comparable: list[Decimal] = []
    weighted: list[tuple[Decimal, Decimal]] = []
    for key in changed:
        old, new = before.get(key), after.get(key)
        if old is None or new is None or old == 0:
            continue
        pct = (new - old) / old * 100
        comparable.append(pct)
        if weights is not None:
            raw_weight = weights.get(key)
            if raw_weight is not None:
                weighted.append((Decimal(str(raw_weight)), pct))

    max_abs = max(abs(pct) for pct in comparable) if comparable else None
    if weighted:
        # An explicit Decimal start keeps the sum Decimal — `sum` starts at int 0,
        # and a Decimal|int mean would not fit RateTableDiff's Decimal field.
        total_weight = sum((weight for weight, _ in weighted), Decimal(0))
        mean = (
            sum((weight * pct for weight, pct in weighted), Decimal(0)) / total_weight
        )
    else:
        mean = None
    return RateTableDiff(
        changed_cells=len(changed),
        max_abs_change_pct=max_abs,
        exposure_weighted_mean_change_pct=mean,
    )


def diff_vs_previous(
    previous_cells: Cells,
    current_cells: Cells,
    keys: Sequence[RateTableKey],
    value: RateTableValue,
    *,
    weights: Weights | None = None,
) -> RateTableDiff:
    """Diff against the immediately prior version (FR-RATE-17)."""
    return _compute_diff(previous_cells, current_cells, keys, value, weights)


def diff_vs_seed(
    seed_cells: Cells,
    current_cells: Cells,
    keys: Sequence[RateTableKey],
    value: RateTableValue,
    *,
    weights: Weights | None = None,
) -> RateTableDiff:
    """Diff against the seed origin, the technical rate (FR-RATE-16/17)."""
    return _compute_diff(seed_cells, current_cells, keys, value, weights)


def _rows_of(table: RateTableVersion) -> list[CellRow]:
    """The version's cells as rows, or a named refusal when storage hides them.

    Parquet-stored versions keep their cells as a blob (FR-RATE-62); the pricing core
    is the pure engine and reads inline rows only, so every operation refuses the blob
    with PARQUET_CELLS_UNAVAILABLE rather than pretending it can open it.
    """
    if table.storage is not RateTableStorageMode.ROWS or table.rows is None:
        raise ValueError(
            f"{PARQUET_CELLS_UNAVAILABLE}: rate_table:{table.slug}@{table.version} "
            "stores its cells as a blob; the pricing core works on inline rows"
        )
    return [{key: str(value) for key, value in row.items()} for row in table.rows]


def _key_domains_of_version(
    table: RateTableVersion, rows: Sequence[CellRow]
) -> dict[str, frozenset[str]]:
    return {key.name: frozenset(row[key.name] for row in rows) for key in table.keys}


def _validate_result(rows: Cells, table: RateTableVersion) -> None:
    """FR-RATE-19 on an operation's result: the first named issue is the refusal."""
    issues = validate_rate_table(
        rows,
        table.keys,
        table.value,
        key_domains=_key_domains_of_version(table, rows),
        default_row=table.default_row,
    )
    if issues:
        raise ValueError(f"{issues[0].code}: {issues[0].message}")


def _ref(table: RateTableVersion, version: int) -> ArtifactRef:
    return ArtifactRef(type="rate_table", slug=table.slug, version=version)


def _new_version(
    table: RateTableVersion,
    rows: list[CellRow],
    change_note: str,
    operation: BulkOperation,
) -> RateTableVersion:
    return table.model_copy(
        update={
            "version": table.version + 1,
            "rows": rows,
            "change_note": change_note,
            "created_by_operation": operation,
        }
    )


def _format_value(value: Decimal) -> str:
    """A decimal string with trailing zeros dropped — never a float (R2)."""
    return str(value.normalize())


def _transform_rows(
    rows: list[CellRow],
    value_name: str,
    transform: Callable[[Decimal], Decimal],
) -> tuple[list[CellRow], int]:
    """Apply the transform per cell, keeping the original string when nothing moves."""
    result: list[CellRow] = []
    changed = 0
    for row in rows:
        old = Decimal(row[value_name])
        new = transform(old)
        if new == old:
            result.append(row)
        else:
            result.append({**row, value_name: _format_value(new)})
            changed += 1
    return result, changed


def _check_filter_key(table: RateTableVersion, filter: KeyFilter) -> None:
    unknown = set(filter) - {key.name for key in table.keys}
    if unknown:
        raise ValueError(
            f"{FILTER_UNKNOWN_KEY}: {sorted(unknown)} are not declared keys of "
            f"rate_table:{table.slug}@{table.version}"
        )


def _row_matches(row: CellRow, filter: KeyFilter) -> bool:
    return all(row[key] in values for key, values in filter.items())


def decide_storage_mode(
    cell_count: int, threshold: int = 250_000
) -> Literal["rows", "parquet"]:
    """FR-RATE-62: rows storage at or below the threshold, parquet above it."""
    return "parquet" if cell_count > threshold else "rows"


def uplift_table(
    table: RateTableVersion, *, percentage: Decimal
) -> RateTableVersion:
    """FR-RATE-18: multiply every cell by (1 + percentage), new immutable version.

    The result is validated against the declared bounds before it persists
    (FR-RATE-19), and the version records the operation (04 §4.4).
    """
    rows = _rows_of(table)
    factor = Decimal(1) + percentage
    new_rows, changed = _transform_rows(rows, table.value.name, lambda old: old * factor)
    _validate_result(new_rows, table)
    operation = UpliftTableOperation(
        kind="uplift_table",
        parameters=UpliftTableParameters(percentage=percentage),
        applied_to=_ref(table, table.version),
        result=BulkOperationResult(
            changed_cells=changed, new_version=_ref(table, table.version + 1)
        ),
    )
    return _new_version(
        table, new_rows, f"uplift_table: percentage={percentage}", operation
    )


def uplift_by_filter(
    table: RateTableVersion, *, percentage: Decimal, filter: KeyFilter
) -> RateTableVersion:
    """FR-RATE-18: uplift only the cells matching the filter, new immutable version."""
    rows = _rows_of(table)
    _check_filter_key(table, filter)
    factor = Decimal(1) + percentage
    value_name = table.value.name
    new_rows: list[CellRow] = []
    changed = 0
    for row in rows:
        if _row_matches(row, filter):
            old = Decimal(row[value_name])
            new = old * factor
            if new == old:
                new_rows.append(row)
            else:
                new_rows.append({**row, value_name: _format_value(new)})
                changed += 1
        else:
            new_rows.append(row)
    _validate_result(new_rows, table)
    operation = UpliftByFilterOperation(
        kind="uplift_by_filter",
        parameters=UpliftByFilterParameters(percentage=percentage, filter=filter),
        applied_to=_ref(table, table.version),
        result=BulkOperationResult(
            changed_cells=changed, new_version=_ref(table, table.version + 1)
        ),
    )
    return _new_version(
        table,
        new_rows,
        f"uplift_by_filter: percentage={percentage} filter={filter}",
        operation,
    )


def floor_and_cap(
    table: RateTableVersion, *, floor: Decimal, cap: Decimal
) -> RateTableVersion:
    """FR-RATE-18: clamp every cell into [floor, cap], new immutable version.

    The floor is enforced at operation time (03 §5.2), not shape-enforced: a version
    may still carry values outside the declared bounds if an earlier save allowed it.
    """
    if floor > cap:
        raise ValueError(f"{FLOOR_ABOVE_CAP}: floor {floor} is above cap {cap}")
    rows = _rows_of(table)
    new_rows, changed = _transform_rows(
        rows, table.value.name, lambda old: max(floor, min(old, cap))
    )
    _validate_result(new_rows, table)
    operation = FloorAndCapOperation(
        kind="floor_and_cap",
        parameters=FloorAndCapParameters(floor=floor, cap=cap),
        applied_to=_ref(table, table.version),
        result=BulkOperationResult(
            changed_cells=changed, new_version=_ref(table, table.version + 1)
        ),
    )
    return _new_version(
        table, new_rows, f"floor_and_cap: floor={floor} cap={cap}", operation
    )


def rebase_to_level(
    table: RateTableVersion, *, base_level: KeyFilter
) -> RateTableVersion:
    """FR-RATE-18: rescale every level against the reference level, new version.

    The reference row's value becomes exactly 1 and everything else is expressed
    relative to it. Every cell is re-stringified (the ratio), so even a row whose
    value equals the reference keeps a normalized rendering.
    """
    rows = _rows_of(table)
    _check_filter_key(table, base_level)
    matching = [row for row in rows if _row_matches(row, base_level)]
    if not matching:
        raise ValueError(f"{REBASE_NO_MATCH}: base_level {base_level} matches no row")
    if len(matching) > 1:
        raise ValueError(
            f"{REBASE_AMBIGUOUS}: base_level {base_level} matches {len(matching)} rows"
        )
    reference = Decimal(matching[0][table.value.name])
    if reference == 0:
        raise ValueError(f"{REBASE_ZERO_REFERENCE}: base_level {base_level} has value 0")
    value_name = table.value.name
    new_rows: list[CellRow] = []
    changed = 0
    for row in rows:
        quotient = Decimal(row[value_name]) / reference
        new_rows.append({**row, value_name: _format_value(quotient)})
        if quotient != Decimal(1):
            changed += 1
    _validate_result(new_rows, table)
    operation = RebaseToLevelOperation(
        kind="rebase_to_level",
        parameters=RebaseToLevelParameters(base_level=base_level),
        applied_to=_ref(table, table.version),
        result=BulkOperationResult(
            changed_cells=changed, new_version=_ref(table, table.version + 1)
        ),
    )
    return _new_version(
        table, new_rows, f"rebase_to_level: base_level={base_level}", operation
    )


def export_to_csv(table: RateTableVersion) -> bytes:
    """FR-RATE-20: the version as CSV — header then rows, decimal strings only."""
    rows = _rows_of(table)
    header = [key.name for key in table.keys] + [table.value.name]
    buffer = StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(header)
    for row in rows:
        writer.writerow([row[column] for column in header])
    return buffer.getvalue().encode("utf-8")


def export_to_xlsx(table: RateTableVersion) -> bytes:
    """FR-RATE-20: the version as XLSX — every cell written as text, never float.

    The strict round-trip (import back to the same decimal strings) requires cells
    that a spreadsheet's number handling cannot silently re-type, so openpyxl cells
    are written as strings (03 §8).
    """
    rows = _rows_of(table)
    header = [key.name for key in table.keys] + [table.value.name]
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append([row[column] for column in header])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _parse_csv(content: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """The file's header and data rows, or a named parse refusal."""
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        raise ValueError(f"{IMPORT_PARSE_ERROR}: {filename} is not valid UTF-8") from None
    rows = [row for row in csv.reader(StringIO(text))]
    if not rows:
        raise ValueError(f"{IMPORT_PARSE_ERROR}: {filename} is empty")
    return rows[0], rows[1:]


def _parse_xlsx(content: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """The file's header and data rows (cells coerced to strings), or a named refusal."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True)
    except (BadZipFile, InvalidFileException):
        raise ValueError(f"{IMPORT_PARSE_ERROR}: {filename} is not a readable XLSX") from None
    rows = [
        ["" if cell is None else str(cell) for cell in row]
        for row in workbook.active.iter_rows(values_only=True)
    ]
    if not rows:
        raise ValueError(f"{IMPORT_PARSE_ERROR}: {filename} is empty")
    return rows[0], rows[1:]


def _value_matches_key_type(raw: str, key_type: RateTableKeyType) -> bool:
    if key_type is RateTableKeyType.INT:
        try:
            int(raw)
        except ValueError:
            return False
        return True
    if key_type is RateTableKeyType.DATE:
        try:
            datetime.fromisoformat(raw)
        except ValueError:
            return False
        return True
    if key_type is RateTableKeyType.BOOL:
        return raw.strip().lower() in {"true", "false"}
    return True


def _check_import_types(table: RateTableVersion, file_rows: list[CellRow]) -> None:
    for key in table.keys:
        if key.type is RateTableKeyType.STRING:
            continue
        for row in file_rows:
            raw = row[key.name]
            if not _value_matches_key_type(raw, key.type):
                raise ValueError(
                    f"{IMPORT_TYPE_MISMATCH}: value {raw!r} for key {key.name} does not "
                    "match its declared key type"
                )


def _import_preview(
    table: RateTableVersion, content: bytes, *, filename: str
) -> ImportPreview:
    """The shared import pipeline (03 §5.2, ruling b): diff and strict verdict.

    The would-be version is checked against the addressed version's own validated
    domain — same keys, same key types, same coverage — so an import can never
    silently drop, re-type, or re-key a cell. Only a confirmed-clean diff creates a
    version, and only the caller (the API) persists it.
    """
    rows = _rows_of(table)
    header, data = (
        _parse_csv(content, filename)
        if filename.endswith(".csv")
        else _parse_xlsx(content, filename)
    )
    expected_header = [key.name for key in table.keys] + [table.value.name]
    if header != expected_header:
        raise ValueError(
            f"{IMPORT_KEY_MISMATCH}: header {header} is not the declared keys and "
            f"value {expected_header}"
        )
    if any(len(row) != len(header) for row in data):
        raise ValueError(f"{IMPORT_PARSE_ERROR}: {filename} has ragged rows")
    file_rows = [dict(zip(header, row, strict=True)) for row in data]
    _check_import_types(table, file_rows)
    issues = validate_rate_table(
        file_rows,
        table.keys,
        table.value,
        key_domains={key.name: frozenset(row[key.name] for row in rows) for key in table.keys},
        default_row=table.default_row,
    )
    if issues:
        raise ValueError(f"{issues[0].code}: {issues[0].message}")
    diff = _compute_diff(rows, file_rows, table.keys, table.value, None)
    verdict = ImportVerdict(
        filename=filename,
        content_sha256=hashlib.sha256(content).hexdigest(),
        round_trip="passed",
    )
    return ImportPreview(diff=diff, created_by_import=verdict)


def import_from_csv(version: RateTableVersion, content: bytes) -> ImportPreview:
    """FR-RATE-20: preview the would-be version against the addressed one (ruling b)."""
    return _import_preview(version, content, filename="import.csv")


def import_from_xlsx(version: RateTableVersion, content: bytes) -> ImportPreview:
    """FR-RATE-20: preview the would-be version against the addressed one (ruling b)."""
    return _import_preview(version, content, filename="import.xlsx")
