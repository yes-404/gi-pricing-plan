"""Bulk operations, storage-mode decision, and CSV/XLSX import/export (W10-3B).

Covers 03 §5.2's rate-table interface beyond the W10-2 surface: the four bulk
operations of FR-233 (each recording its parameters as a BulkOperation on the
version it creates, 04 §4.4), `decide_storage_mode` (FR-232), and the CSV/XLSX
export/import of FR-235 with the strict round-trip verdict. Values are decimal
strings everywhere — never JSON floats, and never float through the file (R2).
"""

from __future__ import annotations

import hashlib
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import ClassVar

import pytest
from openpyxl import load_workbook

from model_schema.rating import (
    ImportPreview,
    ImportVerdict,
    RateTableKey,
    RateTableKeyType,
    RateTableStorageMode,
    RateTableValue,
    RateTableValueType,
    RateTableVersion,
    SeededFrom,
)
from model_schema.refs import ArtifactRef
from pricing_core.rate_tables.operations import (
    decide_storage_mode,
    export_to_csv,
    export_to_xlsx,
    floor_and_cap,
    import_confirmed,
    import_from_csv,
    import_from_xlsx,
    rebase_to_level,
    uplift_by_filter,
    uplift_table,
)

_SLUG = "motor-driver-age-relativity"


def _key(
    name: str = "driver_age_band",
    type: RateTableKeyType = RateTableKeyType.STRING,
) -> RateTableKey:
    return RateTableKey(name=name, type=type, banding_ref=None)


def _value(
    name: str = "relativity", *, min: Decimal | None = Decimal("0.2"),
    max: Decimal | None = Decimal("5.0"),
) -> RateTableValue:
    return RateTableValue(
        name=name, type=RateTableValueType.RELATIVITY, unit="factor", min=min, max=max
    )


#: The §4.2 example's rows: a single key with three complete levels.
_DEFAULT_ROWS = [
    {"driver_age_band": "17-20", "relativity": "1.8400"},
    {"driver_age_band": "21-24", "relativity": "1.4100"},
    {"driver_age_band": "25-29", "relativity": "1.1200"},
]


def _version(
    *,
    version: int = 6,
    rows: list[dict[str, str]] | None = None,
    value: RateTableValue | None = None,
    storage: RateTableStorageMode = RateTableStorageMode.ROWS,
    keys: list[RateTableKey] | None = None,
    seeded_from: SeededFrom | None = None,
) -> RateTableVersion:
    if storage is RateTableStorageMode.PARQUET:
        return RateTableVersion(
            slug=_SLUG,
            version=version,
            rateable=True,
            storage=storage,
            keys=keys if keys is not None else [_key()],
            value=value if value is not None else _value(),
            change_note="parquet test version",
            cells={"sha256": "a" * 64, "bytes": 42, "media_type": "parquet"},
            seeded_from=seeded_from,
        )
    return RateTableVersion(
        slug=_SLUG,
        version=version,
        rateable=True,
        storage=storage,
        keys=keys if keys is not None else [_key()],
        value=value if value is not None else _value(),
        rows=rows if rows is not None else list(_DEFAULT_ROWS),
        change_note="baseline version",
        seeded_from=seeded_from,
    )


class TestDecideStorageMode:
    @pytest.mark.req("FR-232")
    def test_small_table_uses_rows_storage(self) -> None:
        assert decide_storage_mode(100) == "rows"

    @pytest.mark.req("FR-232")
    def test_large_table_uses_parquet_storage(self) -> None:
        assert decide_storage_mode(1_000_000) == "parquet"

    @pytest.mark.req("FR-232")
    def test_at_the_threshold_stays_rows(self) -> None:
        assert decide_storage_mode(250_000) == "rows"

    @pytest.mark.req("FR-232")
    def test_above_the_threshold_spills_to_parquet(self) -> None:
        assert decide_storage_mode(250_001) == "parquet"

    @pytest.mark.req("FR-232")
    def test_honours_the_configured_threshold(self) -> None:
        assert decide_storage_mode(300_000, threshold=100_000) == "parquet"
        assert decide_storage_mode(100_000, threshold=100_000) == "rows"


class TestUpliftTable:
    @pytest.mark.req("FR-233")
    def test_uplifts_every_cell_by_the_percentage(self) -> None:
        result = uplift_table(_version(), percentage=Decimal("0.10"))
        assert result.rows == [
            {"driver_age_band": "17-20", "relativity": "2.024"},
            {"driver_age_band": "21-24", "relativity": "1.551"},
            {"driver_age_band": "25-29", "relativity": "1.232"},
        ]

    @pytest.mark.req("FR-233")
    def test_creates_a_new_immutable_version(self) -> None:
        baseline = _version()
        result = uplift_table(baseline, percentage=Decimal("0.10"))
        assert result.version == 7
        assert result.slug == _SLUG
        assert result.storage is RateTableStorageMode.ROWS
        assert result.rateable is True
        assert result.change_note == "uplift_table: percentage=0.10"
        assert baseline.version == 6
        assert baseline.rows == _DEFAULT_ROWS

    @pytest.mark.req("FR-233")
    def test_records_the_operation_with_parameters_and_result(self) -> None:
        result = uplift_table(_version(), percentage=Decimal("0.10"))
        assert result.created_by_operation is not None
        assert result.created_by_operation.kind == "uplift_table"
        assert result.created_by_operation.parameters.percentage == Decimal("0.10")
        assert str(result.created_by_operation.applied_to) == f"rate_table:{_SLUG}@6"
        assert result.created_by_operation.result.changed_cells == 3
        assert (
            str(result.created_by_operation.result.new_version) == f"rate_table:{_SLUG}@7"
        )

    @pytest.mark.req("FR-233")
    def test_zero_percentage_changes_nothing(self) -> None:
        result = uplift_table(_version(), percentage=Decimal("0"))
        assert result.rows == _DEFAULT_ROWS
        assert result.created_by_operation is not None
        assert result.created_by_operation.result.changed_cells == 0

    @pytest.mark.req("FR-234")
    def test_result_outside_declared_bounds_is_refused(self) -> None:
        with pytest.raises(ValueError, match="OUT_OF_BOUNDS"):
            uplift_table(_version(), percentage=Decimal("2.00"))

    @pytest.mark.req("FR-232")
    def test_parquet_version_is_refused(self) -> None:
        with pytest.raises(ValueError, match="PARQUET_CELLS_UNAVAILABLE"):
            uplift_table(_version(storage=RateTableStorageMode.PARQUET), percentage=Decimal("0.10"))


class TestUpliftByFilter:
    @pytest.mark.req("FR-233")
    def test_uplifts_only_the_matching_cells(self) -> None:
        result = uplift_by_filter(
            _version(),
            percentage=Decimal("0.10"),
            filter={"driver_age_band": ["17-20"]},
        )
        assert result.rows == [
            {"driver_age_band": "17-20", "relativity": "2.024"},
            {"driver_age_band": "21-24", "relativity": "1.4100"},
            {"driver_age_band": "25-29", "relativity": "1.1200"},
        ]
        assert result.created_by_operation is not None
        assert result.created_by_operation.result.changed_cells == 1

    @pytest.mark.req("FR-233")
    def test_filter_matching_nothing_creates_an_unchanged_version(self) -> None:
        result = uplift_by_filter(
            _version(),
            percentage=Decimal("0.10"),
            filter={"driver_age_band": ["60+"]},
        )
        assert result.rows == _DEFAULT_ROWS
        assert result.created_by_operation is not None
        assert result.created_by_operation.result.changed_cells == 0

    @pytest.mark.req("FR-233")
    def test_unknown_filter_key_is_refused(self) -> None:
        with pytest.raises(ValueError, match="FILTER_UNKNOWN_KEY"):
            uplift_by_filter(
                _version(), percentage=Decimal("0.10"), filter={"not_a_key": ["x"]}
            )

    @pytest.mark.req("FR-233")
    def test_filtered_uplift_out_of_bounds_is_refused(self) -> None:
        with pytest.raises(ValueError, match="OUT_OF_BOUNDS"):
            uplift_by_filter(
                _version(), percentage=Decimal("2.00"),
                filter={"driver_age_band": ["17-20"]},
            )


class TestFloorAndCap:
    @pytest.mark.req("FR-233")
    def test_clamps_values_into_the_bounds(self) -> None:
        rows = [
            {"driver_age_band": "17-20", "relativity": "0.3000"},
            {"driver_age_band": "21-24", "relativity": "2.5000"},
            {"driver_age_band": "25-29", "relativity": "1.0000"},
        ]
        result = floor_and_cap(
            _version(rows=rows, value=_value(min=None, max=None)),
            floor=Decimal("0.5"),
            cap=Decimal("2.0"),
        )
        assert result.rows == [
            {"driver_age_band": "17-20", "relativity": "0.5"},
            {"driver_age_band": "21-24", "relativity": "2"},
            {"driver_age_band": "25-29", "relativity": "1.0000"},
        ]
        assert result.created_by_operation is not None
        assert result.created_by_operation.result.changed_cells == 2

    @pytest.mark.req("FR-233")
    def test_unchanged_cells_keep_their_exact_strings(self) -> None:
        result = floor_and_cap(
            _version(), floor=Decimal("0.5"), cap=Decimal("2.0")
        )
        assert result.rows == _DEFAULT_ROWS
        assert result.created_by_operation is not None
        assert result.created_by_operation.result.changed_cells == 0

    @pytest.mark.req("FR-233")
    def test_floor_above_cap_is_refused(self) -> None:
        with pytest.raises(ValueError, match="FLOOR_ABOVE_CAP"):
            floor_and_cap(_version(), floor=Decimal("2.0"), cap=Decimal("0.5"))


class TestRebaseToLevel:
    _ROWS: ClassVar[list[dict[str, str]]] = [
        {"driver_age_band": "17-20", "relativity": "2.0000"},
        {"driver_age_band": "21-24", "relativity": "1.5000"},
        {"driver_age_band": "25-29", "relativity": "1.0000"},
    ]

    @pytest.mark.req("FR-233")
    def test_reference_level_becomes_one_and_everything_rescales(self) -> None:
        result = rebase_to_level(
            _version(rows=self._ROWS, value=_value(min=None, max=None)),
            base_level={"driver_age_band": ["25-29"]},
        )
        assert result.rows == [
            {"driver_age_band": "17-20", "relativity": "2"},
            {"driver_age_band": "21-24", "relativity": "1.5"},
            {"driver_age_band": "25-29", "relativity": "1"},
        ]
        assert result.created_by_operation is not None
        assert result.created_by_operation.parameters.base_level == {
            "driver_age_band": ["25-29"]
        }
        assert result.created_by_operation.result.changed_cells == 2

    @pytest.mark.req("FR-233")
    def test_zero_reference_level_is_refused(self) -> None:
        rows = [
            {"driver_age_band": "17-20", "relativity": "2.0000"},
            {"driver_age_band": "21-24", "relativity": "0.0000"},
            {"driver_age_band": "25-29", "relativity": "1.0000"},
        ]
        with pytest.raises(ValueError, match="REBASE_ZERO_REFERENCE"):
            rebase_to_level(
                _version(rows=rows, value=_value(min=None, max=None)),
                base_level={"driver_age_band": ["21-24"]},
            )

    @pytest.mark.req("FR-233")
    def test_base_level_matching_nothing_is_refused(self) -> None:
        with pytest.raises(ValueError, match="REBASE_NO_MATCH"):
            rebase_to_level(
                _version(rows=self._ROWS, value=_value(min=None, max=None)),
                base_level={"driver_age_band": ["60+"]},
            )

    @pytest.mark.req("FR-233")
    def test_ambiguous_base_level_is_refused(self) -> None:
        with pytest.raises(ValueError, match="REBASE_AMBIGUOUS"):
            rebase_to_level(
                _version(rows=self._ROWS, value=_value(min=None, max=None)),
                base_level={"driver_age_band": ["17-20", "21-24"]},
            )


class TestExportCsv:
    @pytest.mark.req("FR-235")
    def test_exports_header_and_rows_as_decimal_strings(self) -> None:
        content = export_to_csv(_version())
        assert content == (
            b"driver_age_band,relativity\n"
            b"17-20,1.8400\n"
            b"21-24,1.4100\n"
            b"25-29,1.1200\n"
        )

    @pytest.mark.req("FR-232")
    def test_parquet_version_is_refused(self) -> None:
        with pytest.raises(ValueError, match="PARQUET_CELLS_UNAVAILABLE"):
            export_to_csv(_version(storage=RateTableStorageMode.PARQUET))


class TestExportXlsx:
    @pytest.mark.req("FR-235")
    def test_exports_header_and_rows_as_text_never_float(self) -> None:
        content = export_to_xlsx(_version())
        sheet = load_workbook(BytesIO(content), read_only=True).active
        assert [list(row) for row in sheet.iter_rows(values_only=True)] == [
            ["driver_age_band", "relativity"],
            ["17-20", "1.8400"],
            ["21-24", "1.4100"],
            ["25-29", "1.1200"],
        ]
        assert all(
            isinstance(cell, str)
            for row in sheet.iter_rows(values_only=True)
            for cell in row
        )

    @pytest.mark.req("FR-232")
    def test_parquet_version_is_refused(self) -> None:
        with pytest.raises(ValueError, match="PARQUET_CELLS_UNAVAILABLE"):
            export_to_xlsx(_version(storage=RateTableStorageMode.PARQUET))


class TestImport:
    @pytest.mark.req("FR-235")
    def test_round_trip_import_yields_a_passing_verdict(self) -> None:
        content = export_to_csv(_version())
        preview = import_from_csv(_version(), content, filename="rate-change-2026-08.csv")
        assert isinstance(preview, ImportPreview)
        assert preview.diff.changed_cells >= 0
        assert preview.created_by_import.round_trip == "passed"
        assert isinstance(preview.created_by_import, ImportVerdict)
        assert str(preview.created_by_import.applied_to) == f"rate_table:{_SLUG}@6"

    @pytest.mark.req("FR-235")
    def test_xlsx_round_trip_import(self) -> None:
        content = export_to_xlsx(_version())
        preview = import_from_xlsx(_version(), content, filename="rate-change.xlsx")
        assert preview.created_by_import.round_trip == "passed"
        assert str(preview.created_by_import.applied_to) == f"rate_table:{_SLUG}@6"

    @pytest.mark.req("FR-235")
    def test_the_verdict_records_the_upload_name_as_received(self) -> None:
        """DP5: the verdict's filename is the real upload name, not a format constant —
        the only link between the offline artifact and the online audit record."""
        content = export_to_csv(_version())
        preview = import_from_csv(
            _version(), content, filename="2026-08-28-rate-change.csv"
        )
        assert preview.created_by_import.filename == "2026-08-28-rate-change.csv"

    @pytest.mark.req("FR-235")
    def test_the_confirmed_import_returns_the_cells_and_verdict(self) -> None:
        """DP6: `confirm` re-parses the same bytes — the API persists only what a
        strict pass hands it, so preview and created version cannot diverge."""
        content = (
            b"driver_age_band,relativity\n"
            b"17-20,1.8400\n"
            b"21-24,1.4500\n"
            b"25-29,1.1200\n"
        )
        result = import_confirmed(_version(), content, filename="confirmed.csv")
        assert [row["driver_age_band"] for row in result.cells] == [
            "17-20",
            "21-24",
            "25-29",
        ]
        assert [row["relativity"] for row in result.cells] == [
            "1.8400",
            "1.4500",
            "1.1200",
        ]
        assert result.created_by_import.filename == "confirmed.csv"
        assert result.created_by_import.round_trip == "passed"
        assert str(result.created_by_import.applied_to) == f"rate_table:{_SLUG}@6"

    @pytest.mark.req("FR-235")
    def test_the_confirmed_import_refuses_a_verdict_violation(self) -> None:
        """Confirmation cannot override the round-trip verdict: the same named error
        on both calls (DP6)."""
        content = b"vehicle_age_band,relativity\n17-20,1.8400\n"
        with pytest.raises(ValueError, match="IMPORT_KEY_MISMATCH"):
            import_confirmed(_version(), content, filename="wrong-header.csv")

    @pytest.mark.req("FR-235")
    def test_unknown_header_column_is_refused(self) -> None:
        content = b"driver_age_band,relativity,extra\n17-20,1.8400\n"
        with pytest.raises(ValueError, match="IMPORT_KEY_MISMATCH"):
            import_from_csv(_version(), content, filename="import.csv")

    @pytest.mark.req("FR-235")
    def test_missing_key_column_is_refused(self) -> None:
        content = b"relativity\n1.8400\n"
        with pytest.raises(ValueError, match="IMPORT_KEY_MISMATCH"):
            import_from_csv(_version(), content, filename="import.csv")

    @pytest.mark.req("FR-235")
    def test_key_type_mismatch_is_refused(self) -> None:
        content = b"driver_age_band,relativity\n17-20,1.8400\n"
        keys = [_key(type=RateTableKeyType.INT)]
        version = _version(keys=keys)
        with pytest.raises(ValueError, match="IMPORT_TYPE_MISMATCH"):
            import_from_csv(version, content, filename="import.csv")

    @pytest.mark.req("FR-235")
    def test_import_against_a_parquet_version_is_refused(self) -> None:
        content = export_to_csv(_version())
        with pytest.raises(ValueError, match="PARQUET_CELLS_UNAVAILABLE"):
            import_from_csv(
                _version(storage=RateTableStorageMode.PARQUET), content, filename="import.csv"
            )

    @pytest.mark.req("FR-235")
    def test_modified_file_diffs_against_the_addressed_version(self) -> None:
        """Ruling (b) evidenced positively: the diff counts what actually moved."""
        content = (
            b"driver_age_band,relativity\n"
            b"17-20,1.8400\n"
            b"21-24,1.4500\n"
            b"25-29,1.1200\n"
        )
        preview = import_from_csv(_version(), content, filename="import.csv")
        assert preview.diff.changed_cells == 1
        assert (
            preview.created_by_import.content_sha256
            == hashlib.sha256(content).hexdigest()
        )


class TestSeedLineage:
    """The #306 ruling (a): a derived version inherits the baseline's seeded_from."""

    @pytest.mark.req("FR-230")
    def test_derived_version_keeps_the_baseline_seeded_from(self) -> None:
        baseline = _version(
            seeded_from=SeededFrom(
                model_ref=ArtifactRef(type="model", slug="motor", version=3),
                seeded_at=datetime(2026, 8, 28, tzinfo=UTC),
            )
        )
        result = uplift_table(baseline, percentage=Decimal("0.10"))
        assert result.seeded_from is not None
        assert result.seeded_from == baseline.seeded_from

    @pytest.mark.req("FR-230")
    def test_unseeded_baseline_stays_unseeded(self) -> None:
        result = uplift_table(_version(), percentage=Decimal("0.10"))
        assert result.seeded_from is None

    @pytest.mark.req("FR-235")
    def test_missing_key_combination_is_refused(self) -> None:
        keys = [_key("driver_age_band"), _key("region")]
        rows = [
            {"driver_age_band": "17-20", "region": "south", "relativity": "1.8400"},
            {"driver_age_band": "17-20", "region": "north", "relativity": "1.9000"},
            {"driver_age_band": "21-24", "region": "south", "relativity": "1.4100"},
            {"driver_age_band": "21-24", "region": "north", "relativity": "1.4500"},
        ]
        version = _version(keys=keys, rows=rows)
        content = (
            b"driver_age_band,region,relativity\n"
            b"17-20,south,1.8400\n"
            b"17-20,north,1.9000\n"
            b"21-24,south,1.4100\n"
        )
        with pytest.raises(ValueError, match="INCOMPLETE_KEY_DOMAIN"):
            import_from_csv(version, content, filename="import.csv")
