"""Rate table routes (03 §5.1, slices W10-2/W10-3C): seeding, cell diffs, bulk
operations, and the export/import preview.

`POST /rate-tables/{slug}/seed-from-model` (FR-RATE-16),
`GET /rate-tables/{slug}@{version}/diff?against=` (FR-RATE-17),
`POST /rate-tables/{slug}@{version}/bulk-operation` (FR-RATE-18), and the CSV/XLSX
export and import preview (FR-RATE-20). Models are inserted rather than fitted —
these routes care that the model row carries an approved status and a fit result with
relativities, not how the fit happened, and a real GLM fit per test would buy nothing
this file asserts.
"""

from __future__ import annotations

import asyncio
import hashlib
import io
from uuid import UUID, uuid4

import pytest
from backend.tests.test_api_datasets import _headers
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.config import Settings
from app.db.models import (
    ModelRow,
    RateTableRow,
    RateTableVersionRow,
)
from app.db.session import Database
from model_schema import GlmSpec, ModelStatus, OffsetSpec, new_uuid7


@pytest.fixture
def actuary(workspace_id, principal, grant) -> dict[str, str]:
    """A caller with `RATING_WRITE` and `RATING_READ` (the `analyst` role)."""
    asyncio.get_event_loop().run_until_complete(grant("analyst"))
    return _headers(principal.id, workspace_id)


@pytest.fixture
def auditor_headers(workspace_id, grant) -> dict[str, str]:
    """A caller with `RATING_READ` but not `RATING_WRITE` (the `auditor` role)."""
    other = uuid4()
    asyncio.get_event_loop().run_until_complete(grant("auditor", principal_id=other))
    return _headers(other, workspace_id)


@pytest.fixture
def admin_headers(workspace_id, grant) -> dict[str, str]:
    """A caller with `ADMIN_MANAGE_SETTINGS` — the workspace threshold is a setting
    (FR-RATE-62), so the parquet-path tests set it through the API like an operator."""
    other = uuid4()
    asyncio.get_event_loop().run_until_complete(grant("admin", principal_id=other))
    return _headers(other, workspace_id)


def _glm_spec(family: str, dataset_version_id: UUID) -> dict[str, object]:
    """A valid `GlmSpec` as JSON — `to_model` re-validates it, so `{}` will not do."""
    spec = GlmSpec(
        model_family_slug=family,
        dataset_version_id=dataset_version_id,
        response_column="claim_count",
        offset=OffsetSpec(kind="log_column", column="exposure_years"),
    )
    return spec.model_dump(mode="json")


def _fit_result(relativities: dict[str, list[float]]) -> dict[str, object]:
    """A `GlmFitResult` as JSON: one entry per factor, one per level."""
    return {
        "model_type": "glm",
        "converged": True,
        "iterations": 8,
        "fit_seconds": 1.0,
        "relativities": {
            factor: [
                {"level": level, "relativity": relativity, "estimate": 0.5}
                for level, relativity in levels
            ]
            for factor, levels in relativities.items()
        },
    }


def _insert_rows(rows: list[object]) -> None:
    """Insert rows on a loop of our own (`TestClient` is blocking, so an async fixture
    cannot be requested from the synchronous tests below)."""
    from backend.tests.conftest_db import test_database_url

    async def _insert(database: Database) -> None:
        async with database.unit_of_work() as session:
            session.add_all(rows)
            await session.flush()

    loop = asyncio.new_event_loop()
    try:
        database = Database(Settings(database_url=test_database_url()))
        try:
            loop.run_until_complete(_insert(database))
        finally:
            loop.run_until_complete(database.dispose())
    finally:
        loop.close()


def _seed_approved_model(
    workspace_id: UUID, family: str, relativities: dict[str, list[tuple[str, float]]]
) -> None:
    """An approved ModelRow whose fit carries the given relativities."""
    _insert_rows(
        [
            ModelRow(
                workspace_id=workspace_id,
                model_family_slug=family,
                version=1,
                status=ModelStatus.APPROVED.value,
                dataset_version_id=new_uuid7(),
                spec=_glm_spec(family, new_uuid7()),
                spec_hash=f"v3:sha256:{uuid4().hex}{uuid4().hex}",
                fit_result=_fit_result(relativities),
                diagnostics_id=uuid4(),
            )
        ]
    )


_LEVELS: dict[str, list[tuple[str, float]]] = {
    "driver_age_band": [
        ("17-20", 1.92),
        ("21-24", 1.41),
        ("25-29", 1.12),
    ]
}


def _seed_body(family: str, change_note: str = "Seeded for the W10-2 tests") -> dict[str, object]:
    return {"model_ref": f"model:{family}@1", "change_note": change_note}


def _table_slug() -> str:
    return f"motor-driver-age-{uuid4().hex[:8]}"


@pytest.mark.req("FR-RATE-16")
def test_seed_creates_version_one_with_cells(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()

    response = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["slug"] == slug
    assert body["version"] == 1
    assert body["rateable"] is True
    assert body["storage"] == "rows"
    assert [key["name"] for key in body["keys"]] == ["driver_age_band"]
    assert body["value"]["name"] == "relativity"
    assert body["change_note"] == "Seeded for the W10-2 tests"
    assert body["seeded_from"]["model_ref"] == f"model:{family}@1"
    assert [row["driver_age_band"] for row in body["rows"]] == [
        "17-20",
        "21-24",
        "25-29",
    ]
    assert [row["relativity"] for row in body["rows"]] == ["1.92", "1.41", "1.12"]


@pytest.mark.req("FR-RATE-16")
@pytest.mark.req("FR-RATE-17")
def test_seed_appends_the_next_version_and_diff_vs_previous(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()

    first = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert first.status_code == 201, first.text
    assert first.json()["version"] == 1

    softened = {"driver_age_band": [("17-20", 1.84), ("21-24", 1.41), ("25-29", 1.12)]}
    family_v2 = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family_v2, softened)
    second = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family_v2, change_note="Softened 17-20"),
        headers=actuary,
    )
    assert second.status_code == 201, second.text
    assert second.json()["version"] == 2

    diff = api_client.get(
        f"/api/v1/rate-tables/{slug}@2/diff",
        params={"against": "previous"},
        headers=actuary,
    )
    assert diff.status_code == 200, diff.text
    body = diff.json()
    assert body["changed_cells"] == 1
    assert body["max_abs_change_pct"].startswith("4.166")
    assert body["exposure_weighted_mean_change_pct"] is None


@pytest.mark.req("FR-RATE-17")
def test_diff_vs_seed_compares_against_the_origin_not_the_previous_version(
    api_client: TestClient, workspace_id, actuary
) -> None:
    """With three versions, `against=seed` answers v3-vs-v1 where
    `against=previous` answers v3-vs-v2."""
    slug = _table_slug()
    softened = {"driver_age_band": [("17-20", 1.84), ("21-24", 1.41), ("25-29", 1.12)]}
    softened_again = {
        "driver_age_band": [("17-20", 1.84), ("21-24", 1.50), ("25-29", 1.12)]
    }
    for family, relativities in (
        (f"mf-{uuid4().hex[:8]}", _LEVELS),
        (f"mf-{uuid4().hex[:8]}", softened),
        (f"mf-{uuid4().hex[:8]}", softened_again),
    ):
        _seed_approved_model(workspace_id, family, relativities)
        created = api_client.post(
            f"/api/v1/rate-tables/{slug}/seed-from-model",
            json=_seed_body(family),
            headers=actuary,
        )
        assert created.status_code == 201, created.text

    vs_previous = api_client.get(
        f"/api/v1/rate-tables/{slug}@3/diff",
        params={"against": "previous"},
        headers=actuary,
    )
    assert vs_previous.status_code == 200, vs_previous.text
    assert vs_previous.json()["changed_cells"] == 1  # 21-24 only

    vs_seed = api_client.get(
        f"/api/v1/rate-tables/{slug}@3/diff",
        params={"against": "seed"},
        headers=actuary,
    )
    assert vs_seed.status_code == 200, vs_seed.text
    assert vs_seed.json()["changed_cells"] == 2  # 17-20 and 21-24, from the origin


@pytest.mark.req("FR-RATE-17")
def test_diff_against_an_explicit_version(
    api_client: TestClient, workspace_id, actuary
) -> None:
    slug = _table_slug()
    softened = {"driver_age_band": [("17-20", 1.84), ("21-24", 1.41), ("25-29", 1.12)]}
    for family, relativities in (
        (f"mf-{uuid4().hex[:8]}", _LEVELS),
        (f"mf-{uuid4().hex[:8]}", softened),
    ):
        _seed_approved_model(workspace_id, family, relativities)
        created = api_client.post(
            f"/api/v1/rate-tables/{slug}/seed-from-model",
            json=_seed_body(family),
            headers=actuary,
        )
        assert created.status_code == 201, created.text

    diff = api_client.get(
        f"/api/v1/rate-tables/{slug}@2/diff",
        params={"against": "1"},
        headers=actuary,
    )
    assert diff.status_code == 200, diff.text
    assert diff.json()["changed_cells"] == 1


@pytest.mark.req("FR-RATE-16")
def test_seed_refuses_a_non_approved_model(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _insert_rows(
        [
            ModelRow(
                workspace_id=workspace_id,
                model_family_slug=family,
                version=1,
                status=ModelStatus.FITTED.value,
                dataset_version_id=new_uuid7(),
                spec=_glm_spec(family, new_uuid7()),
                spec_hash=f"v3:sha256:{uuid4().hex}{uuid4().hex}",
                fit_result=_fit_result(_LEVELS),
                diagnostics_id=uuid4(),
            )
        ]
    )

    response = api_client.post(
        f"/api/v1/rate-tables/{_table_slug()}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert response.status_code == 422, response.text
    assert response.json()["code"] == "PIN_NOT_APPROVED"


@pytest.mark.req("FR-RATE-16")
def test_seed_refuses_a_missing_model(
    api_client: TestClient, workspace_id, actuary
) -> None:
    response = api_client.post(
        f"/api/v1/rate-tables/{_table_slug()}/seed-from-model",
        json=_seed_body(f"missing-{uuid4().hex[:8]}"),
        headers=actuary,
    )
    assert response.status_code == 404, response.text


@pytest.mark.req("FR-RATE-15")
@pytest.mark.req("FR-RATE-16")
def test_seed_refuses_a_bad_body(api_client: TestClient, workspace_id, actuary) -> None:
    response = api_client.post(
        f"/api/v1/rate-tables/{_table_slug()}/seed-from-model",
        json={"model_ref": "not-a-ref", "change_note": "x"},
        headers=actuary,
    )
    assert response.status_code == 422, response.text
    assert response.json()["code"] == "VALIDATION_FAILED"

    missing_note = api_client.post(
        f"/api/v1/rate-tables/{_table_slug()}/seed-from-model",
        json={"model_ref": "model:whatever@1"},
        headers=actuary,
    )
    assert missing_note.status_code == 422, missing_note.text

    wrong_type = api_client.post(
        f"/api/v1/rate-tables/{_table_slug()}/seed-from-model",
        json={"model_ref": "rating_algorithm:whatever@1", "change_note": "x"},
        headers=actuary,
    )
    assert wrong_type.status_code == 422, wrong_type.text


@pytest.mark.req("FR-RATE-19")
def test_seed_validation_failure_is_named(
    api_client: TestClient, workspace_id, actuary
) -> None:
    """A model whose relativities repeat a level is refused with the named code —
    the plan's DUPLICATE_KEY maps onto 03 §5.2's RATE_TABLE_KEY_DUPLICATE."""
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(
        workspace_id, family, {"driver_age_band": [("17-20", 1.92), ("17-20", 1.50)]}
    )

    response = api_client.post(
        f"/api/v1/rate-tables/{_table_slug()}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert response.status_code == 422, response.text
    assert response.json()["code"] == "RATE_TABLE_KEY_DUPLICATE"


@pytest.mark.req("FR-RATE-17")
def test_diff_404s_for_unknown_table_and_version(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()
    created = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert created.status_code == 201, created.text

    missing_table = api_client.get(
        f"/api/v1/rate-tables/{_table_slug()}@1/diff",
        params={"against": "previous"},
        headers=actuary,
    )
    assert missing_table.status_code == 404, missing_table.text
    assert missing_table.json()["code"] == "RATE_TABLE_MISS"

    missing_version = api_client.get(
        f"/api/v1/rate-tables/{slug}@9/diff",
        params={"against": "previous"},
        headers=actuary,
    )
    assert missing_version.status_code == 404, missing_version.text

    no_previous = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/diff",
        params={"against": "previous"},
        headers=actuary,
    )
    assert no_previous.status_code == 404, no_previous.text


@pytest.mark.req("FR-RATE-17")
def test_diff_rejects_an_unknown_baseline(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()
    created = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert created.status_code == 201, created.text

    response = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/diff",
        params={"against": "banana"},
        headers=actuary,
    )
    assert response.status_code == 422, response.text
    assert response.json()["code"] == "VALIDATION_FAILED"


@pytest.mark.req("FR-RATE-17")
def test_diff_seed_without_a_seed_origin_404s(
    api_client: TestClient, workspace_id, actuary
) -> None:
    """`against=seed` resolves the baseline from the `seeded_from` trail; a version
    carrying no trail (direct inserts — the only creation path in this slice always
    pins a seed origin) answers 404 RATE_TABLE_MISS."""
    slug = _table_slug()
    table_row = RateTableRow(
        workspace_id=workspace_id,
        slug=slug,
        current_version=1,
        created_by=uuid4(),
    )
    _insert_rows([table_row])
    _insert_rows(
        [
            RateTableVersionRow(
                workspace_id=workspace_id,
                rate_table_id=table_row.id,
                version_number=1,
                storage="rows",
                definition={
                    "slug": slug,
                    "version": 1,
                    "rateable": True,
                    "storage": "rows",
                    "keys": [
                        {
                            "name": "driver_age_band",
                            "type": "string",
                            "banding_ref": None,
                        }
                    ],
                    "value": {
                        "name": "relativity",
                        "type": "relativity",
                        "unit": "factor",
                        "min": None,
                        "max": None,
                    },
                    "default_row": None,
                },
                change_note="unseeded probe",
                created_by=uuid4(),
            ),
        ]
    )

    response = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/diff",
        params={"against": "seed"},
        headers=actuary,
    )
    assert response.status_code == 404, response.text
    assert response.json()["code"] == "RATE_TABLE_MISS"


@pytest.mark.req("FR-RATE-62")
def test_a_parquet_version_is_refused_until_w10_3(
    api_client: TestClient, workspace_id, actuary
) -> None:
    """FR-RATE-62's parquet form answers 202 with a Job in W10-3; nothing can yet write
    parquet storage, so a diff touching one is refused with the named code rather than
    fabricating a diff."""
    slug = _table_slug()
    _insert_rows(
        [
            RateTableRow(
                workspace_id=workspace_id,
                slug=slug,
                current_version=1,
                created_by=uuid4(),
            )
        ]
    )

    async def _add_parquet_version(database: Database) -> None:
        async with database.unit_of_work() as session:
            table_row = (
                await session.execute(
                    select(RateTableRow).where(
                        RateTableRow.workspace_id == workspace_id,
                        RateTableRow.slug == slug,
                    )
                )
            ).scalar_one()
            session.add(
                RateTableVersionRow(
                    workspace_id=workspace_id,
                    rate_table_id=table_row.id,
                    version_number=1,
                    storage="parquet",
                    definition={
                        "slug": slug,
                        "version": 1,
                        "rateable": True,
                        "storage": "parquet",
                        "keys": [
                            {
                                "name": "driver_age_band",
                                "type": "string",
                                "banding_ref": None,
                            }
                        ],
                        "value": {
                            "name": "relativity",
                            "type": "relativity",
                            "unit": "factor",
                            "min": None,
                            "max": None,
                        },
                        "default_row": None,
                    },
                    change_note="parquet probe",
                    created_by=uuid4(),
                )
            )
            await session.flush()

    from backend.tests.conftest_db import test_database_url

    loop = asyncio.new_event_loop()
    try:
        database = Database(Settings(database_url=test_database_url()))
        try:
            loop.run_until_complete(_add_parquet_version(database))
        finally:
            loop.run_until_complete(database.dispose())
    finally:
        loop.close()

    response = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/diff",
        params={"against": "previous"},
        headers=actuary,
    )
    assert response.status_code == 501, response.text
    assert response.json()["code"] == "RATE_TABLE_PARQUET_UNBUILT"


@pytest.mark.req("FR-PLAT-47")
def test_routes_are_permission_gated(
    api_client: TestClient, workspace_id, principal, auditor_headers, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()

    anon_seed = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
    )
    assert anon_seed.status_code == 401, anon_seed.text

    read_only_seed = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=auditor_headers,
    )
    assert read_only_seed.status_code == 403, read_only_seed.text

    created = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert created.status_code == 201, created.text

    anon_diff = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/diff",
        params={"against": "previous"},
    )
    assert anon_diff.status_code == 401, anon_diff.text

    read_only_diff = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/diff",
        # Version 1 has no `previous` (that diff is a 404, asserted in the 404s test);
        # the seed origin is version 1 itself, so this answers 200 with zero changes.
        params={"against": "seed"},
        headers=auditor_headers,
    )
    assert read_only_diff.status_code == 200, read_only_diff.text

    anon_operation = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/bulk-operation",
        json=_bulk_body("uplift_table", {"percentage": "0.10"}),
    )
    assert anon_operation.status_code == 401, anon_operation.text

    read_only_operation = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/bulk-operation",
        json=_bulk_body("uplift_table", {"percentage": "0.10"}),
        headers=auditor_headers,
    )
    assert read_only_operation.status_code == 403, read_only_operation.text

    anon_export = api_client.get(f"/api/v1/rate-tables/{slug}@1/export/csv")
    assert anon_export.status_code == 401, anon_export.text

    read_only_export = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/export/csv",
        headers=auditor_headers,
    )
    assert read_only_export.status_code == 200, read_only_export.text

    anon_import = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/import",
        files={"file": ("import.csv", b"driver_age_band,relativity\n17-20,1.92\n", "text/csv")},
    )
    assert anon_import.status_code == 401, anon_import.text

    read_only_import = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/import",
        files={"file": ("import.csv", b"driver_age_band,relativity\n17-20,1.92\n", "text/csv")},
        headers=auditor_headers,
    )
    assert read_only_import.status_code == 403, read_only_import.text


def _bulk_body(kind: str, parameters: dict[str, object]) -> dict[str, object]:
    return {"kind": kind, "parameters": parameters}


@pytest.mark.req("FR-RATE-18")
def test_bulk_operation_creates_a_new_version_with_the_operation_record(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()
    seeded = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert seeded.status_code == 201, seeded.text

    response = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/bulk-operation",
        json=_bulk_body("uplift_table", {"percentage": "0.10"}),
        headers=actuary,
    )

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["version"] == 2
    assert body["storage"] == "rows"
    assert [row["relativity"] for row in body["rows"]] == ["2.112", "1.551", "1.232"]
    operation = body["created_by_operation"]
    assert operation["kind"] == "uplift_table"
    assert operation["parameters"] == {"percentage": "0.10"}
    assert operation["applied_to"] == f"rate_table:{slug}@1"
    assert operation["result"] == {
        "changed_cells": 3,
        "new_version": f"rate_table:{slug}@2",
    }
    assert body["created_by_import"] is None
    assert body["seeded_from"]["model_ref"] == f"model:{family}@1"


@pytest.mark.req("FR-RATE-18")
def test_bulk_operation_refuses_an_unknown_kind(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()
    seeded = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert seeded.status_code == 201, seeded.text

    response = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/bulk-operation",
        json=_bulk_body("lift_everything", {}),
        headers=actuary,
    )
    assert response.status_code == 422, response.text
    assert response.json()["code"] == "VALIDATION_FAILED"


@pytest.mark.req("FR-RATE-18")
def test_bulk_operation_refuses_floor_above_cap(
    api_client: TestClient, workspace_id, actuary
) -> None:
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()
    seeded = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert seeded.status_code == 201, seeded.text

    response = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/bulk-operation",
        json=_bulk_body("floor_and_cap", {"floor": "2.0", "cap": "1.0"}),
        headers=actuary,
    )
    assert response.status_code == 422, response.text
    assert response.json()["code"] == "FLOOR_ABOVE_CAP"


@pytest.mark.req("FR-RATE-18")
def test_bulk_operation_on_a_missing_version_is_refused(
    api_client: TestClient, workspace_id, actuary
) -> None:
    slug = _table_slug()
    response = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/bulk-operation",
        json=_bulk_body("uplift_table", {"percentage": "0.10"}),
        headers=actuary,
    )
    assert response.status_code == 404, response.text
    assert response.json()["code"] == "RATE_TABLE_MISS"


def _set_threshold(
    api_client: TestClient, admin_headers: dict[str, str], value: int
) -> None:
    """The workspace cell-count threshold through the settings API (FR-RATE-62)."""
    response = api_client.put(
        "/api/v1/settings",
        json={"values": {"rate_tables.cell_threshold": value}},
        headers=admin_headers,
    )
    assert response.status_code == 200, response.text


def _seeded_table(
    api_client: TestClient, workspace_id, actuary
) -> tuple[str, str]:
    """Seed a table and return (slug, family) — the W10-3C test shorthand."""
    family = f"mf-{uuid4().hex[:8]}"
    _seed_approved_model(workspace_id, family, _LEVELS)
    slug = _table_slug()
    seeded = api_client.post(
        f"/api/v1/rate-tables/{slug}/seed-from-model",
        json=_seed_body(family),
        headers=actuary,
    )
    assert seeded.status_code == 201, seeded.text
    return slug, family


@pytest.mark.req("FR-RATE-20")
def test_export_csv_returns_the_seeded_cells(
    api_client: TestClient, workspace_id, actuary
) -> None:
    slug, _ = _seeded_table(api_client, workspace_id, actuary)

    response = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/export/csv",
        headers=actuary,
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith("text/csv")
    assert response.content == (
        b"driver_age_band,relativity\n"
        b"17-20,1.92\n"
        b"21-24,1.41\n"
        b"25-29,1.12\n"
    )


@pytest.mark.req("FR-RATE-20")
def test_export_xlsx_parses_to_the_seeded_cells(
    api_client: TestClient, workspace_id, actuary
) -> None:
    slug, _ = _seeded_table(api_client, workspace_id, actuary)

    response = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/export/xlsx",
        headers=actuary,
    )

    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    sheet = workbook.active
    assert [row for row in sheet.iter_rows(values_only=True)] == [
        ("driver_age_band", "relativity"),
        ("17-20", "1.92"),
        ("21-24", "1.41"),
        ("25-29", "1.12"),
    ]


@pytest.mark.req("FR-RATE-20")
def test_import_previews_a_modified_export(
    api_client: TestClient, workspace_id, actuary
) -> None:
    slug, _ = _seeded_table(api_client, workspace_id, actuary)
    exported = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/export/csv",
        headers=actuary,
    )
    assert exported.status_code == 200, exported.text
    modified = exported.content.replace(b"21-24,1.41", b"21-24,1.4500")

    response = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/import",
        files={"file": ("import.csv", modified, "text/csv")},
        headers=actuary,
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["diff"]["changed_cells"] == 1
    verdict = body["created_by_import"]
    assert verdict["content_sha256"] == hashlib.sha256(modified).hexdigest()
    assert verdict["round_trip"] == "passed"
    assert verdict["applied_to"] == f"rate_table:{slug}@1"


@pytest.mark.req("FR-RATE-20")
def test_import_refuses_a_wrong_header(
    api_client: TestClient, workspace_id, actuary
) -> None:
    slug, _ = _seeded_table(api_client, workspace_id, actuary)

    response = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/import",
        files={"file": ("import.csv", b"vehicle_age_band,relativity\n17-20,1.92\n", "text/csv")},
        headers=actuary,
    )

    assert response.status_code == 422, response.text
    assert response.json()["code"] == "IMPORT_KEY_MISMATCH"


@pytest.mark.req("FR-RATE-62")
@pytest.mark.req("FR-RATE-20")
def test_export_reads_a_parquet_version_inline(
    api_client: TestClient, workspace_id, actuary, admin_headers
) -> None:
    """Above the threshold the version lives as a parquet blob; export materialises
    the cells from it — the Job-worthy read is the diff (W10-3D), not an export."""
    _set_threshold(api_client, admin_headers, 2)
    slug, _ = _seeded_table(api_client, workspace_id, actuary)

    response = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/export/csv",
        headers=actuary,
    )

    assert response.status_code == 200, response.text
    assert response.content == (
        b"driver_age_band,relativity\n"
        b"17-20,1.92\n"
        b"21-24,1.41\n"
        b"25-29,1.12\n"
    )


@pytest.mark.req("FR-RATE-62")
@pytest.mark.req("FR-RATE-20")
def test_import_diffs_against_a_parquet_baseline(
    api_client: TestClient, workspace_id, actuary, admin_headers
) -> None:
    _set_threshold(api_client, admin_headers, 2)
    slug, _ = _seeded_table(api_client, workspace_id, actuary)
    exported = api_client.get(
        f"/api/v1/rate-tables/{slug}@1/export/csv",
        headers=actuary,
    )
    assert exported.status_code == 200, exported.text
    modified = exported.content.replace(b"21-24,1.41", b"21-24,1.4500")

    response = api_client.post(
        f"/api/v1/rate-tables/{slug}@1/import",
        files={"file": ("import.csv", modified, "text/csv")},
        headers=actuary,
    )

    assert response.status_code == 200, response.text
    assert response.json()["diff"]["changed_cells"] == 1
